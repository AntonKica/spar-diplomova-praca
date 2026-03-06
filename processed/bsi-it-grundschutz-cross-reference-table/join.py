import openpyxl
filename = "bsi_it_gs_comp_2022_krt.xlsx"

wb_res = openpyxl.Workbook()
ws_res = wb_res.active
ws_res.cell(row=1, column=1, value="module name")
ws_res.cell(row=1, column=2, value="elementary threats")


wb = openpyxl.load_workbook(filename, read_only=True, data_only=True)
for module_index, module_ws in enumerate(wb.worksheets):
    module_name = module_ws.title.removeprefix("KRT_").removesuffix(".xlsx").replace(".", "_")

    threats = []
    threat_index = 0
    while True:
        cell_value = module_ws.cell(row=1, column=threat_index + 4).value
        if cell_value is None:
            break

        threats.append(cell_value.replace(".", "_").replace(" ", "_"))
        threat_index += 1
    ws_res.cell(row=module_index + 2, column=1, value=module_name)
    ws_res.cell(row=module_index + 2, column=2, value=", ".join(threats))

wb_res.save("module-threat.xlsx")
